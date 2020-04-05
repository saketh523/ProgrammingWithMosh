# SpreadSheet Processor program to update all the transaction values with a discount of 10%
# Also displaying a Barchart on xls file using openpyxl.chart

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
  wb = xl.load_workbook(filename)
  sheet = wb['Sheet1']

  #iterate over all rows having data in the xlsx doc
  for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

  #Getting a BarChart data and adding it to the chart
  values =  Reference(sheet, 
            min_row = 2, 
            max_row = sheet. max_row,
            min_col = 4,
            max_col = 4)

  chart = BarChart()
  chart.add_data(values)
  sheet.add_chart(chart,'e2')

  wb.save(filename)